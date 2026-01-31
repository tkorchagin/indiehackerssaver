import os
import re
import sys
import logging
import asyncio
import random
import subprocess
import requests
import trafilatura
import feedparser
from pathlib import Path
from dotenv import load_dotenv
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.constants import ChatAction
from telegram.ext import ApplicationBuilder, ContextTypes, MessageHandler, CommandHandler, CallbackQueryHandler, filters
from litellm import completion
from openpyxl import load_workbook
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger
import pytz
from html2image import Html2Image
import base64

# Fun status messages while processing
STATUS_MESSAGES = [
    "🧠 Читаю статью...",
    "🔍 Ищу инсайты...",
    "✍️ Пишу черновик...",
    "🎨 Добавляю огонька...",
    "💡 Формулирую мысль...",
    "🔥 Делаю пост виральным...",
    "📝 Редактирую текст...",
    "🎯 Ловлю суть...",
    "⚡ Генерирую контент...",
    "💭 Анализирую историю...",
    "🏗️ Собираю пост...",
    "✨ Полирую формулировки...",
]

# Maximum number of stored generations per user
MAX_STORED_GENERATIONS = 10

# Message batching settings
MESSAGE_BATCH_DELAY = 0.3  # seconds to wait for more messages
message_batches = {}  # user_id -> {messages: [], first_update: Update, task: Task}

# Prompt modifiers for regeneration
PROMPT_MODIFIERS = {
    "shorter": """ВАЖНО: Сделай пост ОЧЕНЬ коротким. Максимум 500 символов.
Оставь только самую суть, одну главную мысль. Убери всё лишнее.""",

    "regenerate": """ВАЖНО: Напиши ДРУГОЙ вариант поста с ДРУГИМ углом подачи.
Используй другую структуру, другой хук, другие акценты. Не повторяй предыдущий вариант.""",

    "no_promo": """ВАЖНО: Убери из поста ВСЕ приглашения, призывы к действию, ссылки на курсы, подписки, каналы и любые промо-элементы.
Оставь только основной контент без призывов что-то купить, подписаться или перейти куда-то. Пост должен быть чисто информативным.""",

    "custom": """ВАЖНО: Учти эти правки при переписывании поста:
{edits}"""
}

# Load environment variables
load_dotenv()

TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")
OPENROUTER_API_KEY = os.getenv("OPENROUTER_API_KEY")
LLM_MODEL = os.getenv("LLM_MODEL", "openrouter/google/gemini-3-pro-preview")
ADMIN_ID = os.getenv("ADMIN_ID")

# Proxy URL for fallback content fetching
PROXY_URL = "https://api.allorigins.win/raw?url="

# Configure logging
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

def extract_url_with_parens(text, start_pos):
    """Extract URL starting at start_pos, handling balanced parentheses (for Wikipedia URLs)."""
    url = ""
    paren_depth = 0
    i = start_pos

    while i < len(text):
        char = text[i]
        # Stop at whitespace or certain characters
        if char in ' \t\n\r<>"\'':
            break
        if char == '(':
            paren_depth += 1
            url += char
        elif char == ')':
            if paren_depth > 0:
                paren_depth -= 1
                url += char
            else:
                # Unmatched ) - stop here (likely markdown syntax)
                break
        else:
            url += char
        i += 1

    return url


def strip_utm_params(text):
    """Remove UTM parameters from all URLs in text."""
    from urllib.parse import urlparse, parse_qs, urlencode, urlunparse

    def clean_single_url(url):
        """Clean UTM params from a single URL."""
        try:
            parsed = urlparse(url)
            params = parse_qs(parsed.query, keep_blank_values=True)
            clean_params = {k: v for k, v in params.items() if not k.lower().startswith('utm_')}
            clean_query = urlencode(clean_params, doseq=True)
            result = urlunparse((
                parsed.scheme,
                parsed.netloc,
                parsed.path,
                parsed.params,
                clean_query,
                parsed.fragment
            ))
            if result.endswith('?'):
                result = result[:-1]
            return result
        except:
            return url

    # Find and replace URLs with balanced parentheses handling
    result = []
    i = 0
    while i < len(text):
        # Look for http:// or https://
        if text[i:i+7].lower() == 'http://' or text[i:i+8].lower() == 'https://':
            url = extract_url_with_parens(text, i)
            clean_url = clean_single_url(url)
            result.append(clean_url)
            i += len(url)
        else:
            result.append(text[i])
            i += 1

    return ''.join(result)


def markdown_to_html(text):
    """Convert Markdown formatting to Telegram HTML."""
    import html

    # Replace guillemets (« ») with regular quotes
    text = text.replace('«', '"').replace('»', '"')

    # Replace ё with е
    text = text.replace('ё', 'е').replace('Ё', 'Е')

    # First, protect URLs from being modified
    url_pattern = r'(https?://[^\s]+)'
    urls = re.findall(url_pattern, text)
    url_placeholders = {}
    for i, url in enumerate(urls):
        placeholder = f"__URL_PLACEHOLDER_{i}__"
        url_placeholders[placeholder] = html.escape(url)
        text = text.replace(url, placeholder)

    # Escape HTML special characters (except our placeholders)
    text = html.escape(text)

    # Restore URL placeholders
    for placeholder, url in url_placeholders.items():
        text = text.replace(html.escape(placeholder), url)

    # Convert **bold** to <b>bold</b>
    text = re.sub(r'\*\*(.+?)\*\*', r'<b>\1</b>', text)

    # Convert *italic* to <i>italic</i> (but not inside URLs)
    text = re.sub(r'(?<![/:])\*([^*\n]+?)\*(?![/])', r'<i>\1</i>', text)

    # Convert `code` to <code>code</code>
    text = re.sub(r'`([^`]+?)`', r'<code>\1</code>', text)

    # Convert [text](url) to <a href="url">text</a>
    text = re.sub(r'\[([^\]]+)\]\(([^)]+)\)', r'<a href="\2">\1</a>', text)

    return text

def fetch_via_proxy(url):
    """Fetch URL content via proxy service."""
    try:
        proxy_url = f"{PROXY_URL}{url}"
        logger.info(f"Trying proxy fetch: {proxy_url}")
        response = requests.get(proxy_url, timeout=30)
        if response.status_code == 200 and len(response.text) > 100:
            logger.info(f"Proxy fetch successful ({len(response.text)} chars)")
            return response.text
        logger.warning(f"Proxy returned status {response.status_code} or short content")
    except Exception as e:
        logger.warning(f"Proxy fetch failed: {e}")
    return None

def extract_content(text):
    """Extract content from URL or return original text. Returns None if URL extraction fails."""
    url_pattern = r'https?://[^\s]+'
    urls = re.findall(url_pattern, text)

    # If text is substantial (>300 chars), use it as-is even if it contains URLs
    if len(text) > 300:
        logger.info(f"Using provided text as content ({len(text)} chars)")
        return text

    if urls:
        url = urls[0]
        logger.info(f"Extracting content from URL: {url}")

        # Try direct fetch first
        downloaded = trafilatura.fetch_url(url)
        if downloaded:
            extracted = trafilatura.extract(downloaded, output_format='markdown', include_links=True)
            if extracted and len(extracted) > 100:
                logger.info(f"Successfully extracted {len(extracted)} chars from {url}")
                return extracted
            logger.warning(f"Trafilatura returned empty/short content for {url}")
        else:
            logger.warning(f"Direct fetch failed for {url}")

        # Fallback to proxy
        logger.info(f"Trying proxy fallback for {url}")
        proxy_html = fetch_via_proxy(url)
        if proxy_html:
            extracted = trafilatura.extract(proxy_html, output_format='markdown', include_links=True)
            if extracted and len(extracted) > 100:
                logger.info(f"Successfully extracted {len(extracted)} chars via proxy from {url}")
                return extracted
            logger.warning(f"Proxy HTML extraction returned empty/short content for {url}")

        return None
    return text

def get_prompt():
    """Read prompt from prompt.md file."""
    try:
        with open("prompt.md", "r", encoding="utf-8") as f:
            content = f.read().strip()
            logger.info(f"Loaded prompt from prompt.md ({len(content)} chars)")
            return content
    except Exception as e:
        logger.error(f"Error reading prompt.md: {e}")
        return "You are a helpful assistant."

def create_post_keyboard(message_id: int) -> InlineKeyboardMarkup:
    """Create inline keyboard with post modification buttons."""
    keyboard = [
        [
            InlineKeyboardButton("✂️ Короче", callback_data=f"shorter:{message_id}"),
            InlineKeyboardButton("🔄 Переделать", callback_data=f"regenerate:{message_id}"),
        ],
        [
            InlineKeyboardButton("🚫 Убрать промо", callback_data=f"no_promo:{message_id}"),
        ]
    ]
    return InlineKeyboardMarkup(keyboard)

def save_generation(context: ContextTypes.DEFAULT_TYPE, message_id: int, original_content: str, generated_post: str):
    """Save generation data to user_data for later regeneration."""
    if 'generations' not in context.user_data:
        context.user_data['generations'] = {}

    # Store generation data
    context.user_data['generations'][str(message_id)] = {
        'original_content': original_content,
        'generated_post': generated_post,
    }

    # Clean up old generations if too many
    generations = context.user_data['generations']
    if len(generations) > MAX_STORED_GENERATIONS:
        # Remove oldest entries (smallest message_ids)
        sorted_ids = sorted(generations.keys(), key=int)
        for old_id in sorted_ids[:len(generations) - MAX_STORED_GENERATIONS]:
            del generations[old_id]

def get_generation(context: ContextTypes.DEFAULT_TYPE, message_id: int) -> dict | None:
    """Get saved generation data by message_id."""
    if 'generations' not in context.user_data:
        return None
    return context.user_data['generations'].get(str(message_id))

async def update_status_periodically(status_message, stop_event):
    """Update status message with fun phrases until stopped."""
    messages = STATUS_MESSAGES.copy()
    random.shuffle(messages)
    index = 0

    while not stop_event.is_set():
        await asyncio.sleep(2.5)
        if stop_event.is_set():
            break
        try:
            await status_message.edit_text(messages[index % len(messages)])
            index += 1
        except Exception:
            pass

async def regenerate_post(update: Update, context: ContextTypes.DEFAULT_TYPE, gen_data: dict, modifier: str, custom_edits: str = None):
    """Regenerate post with a modifier."""
    chat_id = update.effective_chat.id

    # Send status message
    await context.bot.send_chat_action(chat_id=chat_id, action=ChatAction.TYPING)
    status_message = await context.bot.send_message(chat_id=chat_id, text="🔄 Перегенерирую пост...")

    # Start status updater
    stop_event = asyncio.Event()
    status_task = asyncio.create_task(update_status_periodically(status_message, stop_event))

    try:
        system_prompt = get_prompt()

        # Add modifier to the prompt
        if modifier == "custom" and custom_edits:
            modifier_text = PROMPT_MODIFIERS["custom"].format(edits=custom_edits)
        else:
            modifier_text = PROMPT_MODIFIERS.get(modifier, "")

        # Include the previous generated post for context
        user_content = f"""Оригинальный контент:
{gen_data['original_content']}

---
Предыдущий сгенерированный пост:
{gen_data['generated_post']}

---
{modifier_text}"""

        logger.info(f"Regenerating post with modifier: {modifier}")
        response = await asyncio.to_thread(
            completion,
            model=LLM_MODEL,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_content}
            ],
            api_key=OPENROUTER_API_KEY
        )

        reply_text = response.choices[0].message.content
        logger.info(f"Regenerated post ({len(reply_text)} chars)")

        # Stop status updater
        stop_event.set()
        status_task.cancel()

        # Delete status and send new post
        await status_message.delete()
        reply_text = strip_utm_params(reply_text)

        try:
            html_text = markdown_to_html(reply_text)
            sent_message = await context.bot.send_message(
                chat_id=chat_id,
                text=html_text,
                parse_mode='HTML',
                reply_markup=create_post_keyboard(0)  # Placeholder, will update
            )
        except Exception as parse_error:
            logger.warning(f"HTML parse failed, sending as plain text: {parse_error}")
            sent_message = await context.bot.send_message(
                chat_id=chat_id,
                text=reply_text,
                reply_markup=create_post_keyboard(0)
            )

        # Update keyboard with correct message_id and save generation
        await sent_message.edit_reply_markup(reply_markup=create_post_keyboard(sent_message.message_id))
        save_generation(context, sent_message.message_id, gen_data['original_content'], reply_text)

        return sent_message

    except Exception as e:
        stop_event.set()
        status_task.cancel()
        logger.error(f"Error in regenerate_post: {str(e)}", exc_info=True)
        await status_message.edit_text(f"❌ Ошибка при перегенерации: {str(e)[:200]}")
        return None

async def handle_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle inline keyboard button presses."""
    query = update.callback_query
    await query.answer()

    user_id = str(query.from_user.id)

    # Filter by Admin ID if set
    if ADMIN_ID and user_id != str(ADMIN_ID):
        logger.warning(f"Unauthorized callback attempt by {user_id}")
        return

    data = query.data
    action, message_id = data.split(":", 1)
    logger.info(f"Callback: action={action}, message_id={message_id}")

    # Get generation data
    gen_data = get_generation(context, int(message_id))
    if not gen_data:
        await query.message.reply_text("❌ Данные о генерации не найдены. Попробуйте сгенерировать пост заново.")
        return

    if action == "shorter":
        await regenerate_post(update, context, gen_data, "shorter")
    elif action == "regenerate":
        await regenerate_post(update, context, gen_data, "regenerate")
    elif action == "no_promo":
        await regenerate_post(update, context, gen_data, "no_promo")
    elif action == "edits":
        # Set waiting for edits state
        context.user_data['waiting_for_edits'] = int(message_id)
        await query.message.reply_text("✏️ Напишите, какие правки внести в пост:")

async def process_message_content(update: Update, context: ContextTypes.DEFAULT_TYPE, user_text: str):
    """Main message processing logic. Called after message batching."""
    user = update.effective_user
    user_id = str(user.id)

    # Send typing action and status message
    await context.bot.send_chat_action(chat_id=update.effective_chat.id, action=ChatAction.TYPING)
    status_message = await update.message.reply_text("🧠 Читаю статью...")

    # Start status updater
    stop_event = asyncio.Event()
    status_task = asyncio.create_task(update_status_periodically(status_message, stop_event))

    try:
        # Extract content
        content = extract_content(user_text)

        # Check if extraction failed
        if content is None:
            stop_event.set()
            status_task.cancel()
            await status_message.edit_text("❌ Не удалось извлечь контент из ссылки. Сайт может блокировать парсинг или использовать JS-рендеринг.")
            return

        system_prompt = get_prompt()

        # Save extracted content to debug file
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        debug_dir = Path("debug")
        debug_dir.mkdir(exist_ok=True)

        extracted_file = debug_dir / f"{user_id}_{timestamp}_extracted.md"
        with open(extracted_file, "w", encoding="utf-8") as f:
            f.write(content)
        logger.info(f"Saved extracted content to {extracted_file}")

        # Call LLM in thread to not block
        logger.info(f"Calling LLM ({LLM_MODEL})...")
        response = await asyncio.to_thread(
            completion,
            model=LLM_MODEL,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": content}
            ],
            api_key=OPENROUTER_API_KEY
        )

        reply_text = response.choices[0].message.content
        logger.info(f"Successfully received response from LLM ({len(reply_text)} chars)")
        logger.info(f"LLM Response:\n{reply_text}")

        # Stop status updater
        stop_event.set()
        status_task.cancel()

        # Save LLM response to debug file
        response_file = debug_dir / f"{user_id}_{timestamp}_response.md"
        with open(response_file, "w", encoding="utf-8") as f:
            f.write(reply_text)
        logger.info(f"Saved LLM response to {response_file}")

        # Delete status message and send response as reply to original message with keyboard
        await status_message.delete()
        reply_text = strip_utm_params(reply_text)
        try:
            html_text = markdown_to_html(reply_text)
            sent_message = await update.message.reply_text(
                html_text,
                parse_mode='HTML',
                do_quote=True,
                reply_markup=create_post_keyboard(0)  # Placeholder, will update
            )
        except Exception as parse_error:
            logger.warning(f"HTML parse failed, sending as plain text: {parse_error}")
            sent_message = await update.message.reply_text(
                reply_text,
                do_quote=True,
                reply_markup=create_post_keyboard(0)
            )

        # Update keyboard with correct message_id and save generation
        await sent_message.edit_reply_markup(reply_markup=create_post_keyboard(sent_message.message_id))
        save_generation(context, sent_message.message_id, content, reply_text)

    except Exception as e:
        # Stop status updater
        stop_event.set()
        status_task.cancel()

        logger.error(f"Error in process_message_content: {str(e)}", exc_info=True)
        error_msg = str(e)
        if "quota" in error_msg.lower() or "429" in error_msg:
            friendly_error = "🔔 Лимит запросов исчерпан (Quota Exceeded). Пожалуйста, подождите или проверьте биллинг в OpenRouter."
        else:
            friendly_error = f"❌ Произошла ошибка: {error_msg[:200]}"

        await status_message.edit_text(friendly_error)


async def process_batched_messages(user_id: str, context: ContextTypes.DEFAULT_TYPE):
    """Process accumulated messages after batch delay."""
    await asyncio.sleep(MESSAGE_BATCH_DELAY)

    if user_id not in message_batches:
        return

    batch = message_batches.pop(user_id)
    messages = batch['messages']
    first_update = batch['first_update']
    replied_message_id = batch.get('replied_message_id')
    waiting_for_edits_id = batch.get('waiting_for_edits_id')

    # Concatenate all messages
    combined_text = '\n\n'.join(messages)
    logger.info(f"Processing batch of {len(messages)} messages for user {user_id}: {combined_text[:100]}...")

    # Check if user was waiting for custom edits
    if waiting_for_edits_id:
        gen_data = get_generation(context, waiting_for_edits_id)
        if gen_data:
            await regenerate_post(first_update, context, gen_data, "custom", custom_edits=combined_text)
            return

    # Check if this is a reply to a generated post
    if replied_message_id:
        gen_data = get_generation(context, replied_message_id)
        if gen_data:
            # Check for reply commands in the combined text
            text_lower = combined_text.lower().strip()
            if text_lower in ['короче', 'shorter']:
                await regenerate_post(first_update, context, gen_data, "shorter")
                return
            elif text_lower in ['другой', 'еще', 'ещe', 'еще раз', 'другой вариант']:
                await regenerate_post(first_update, context, gen_data, "regenerate")
                return
            else:
                # Treat as custom edits with combined text
                await regenerate_post(first_update, context, gen_data, "custom", custom_edits=combined_text)
                return

    # Process the combined message using the first update for reply
    await process_message_content(first_update, context, combined_text)


async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    user_id = str(user.id)
    user_text = update.message.text

    logger.info(f"Received message from {user.username} ({user_id}): {user_text[:50]}...")

    # Filter by Admin ID if set
    if ADMIN_ID and user_id != str(ADMIN_ID):
        logger.warning(f"Unauthorized access attempt by {user_id}")
        return

    if not user_text:
        return

    # Check if user is providing custom edits after pressing "📝 Правки" button
    if context.user_data.get('waiting_for_edits'):
        message_id = context.user_data.pop('waiting_for_edits')
        gen_data = get_generation(context, message_id)
        if gen_data:
            # Still use batching for edits
            if user_id in message_batches:
                batch = message_batches[user_id]
                batch['task'].cancel()
                batch['messages'].append(user_text)
                batch['waiting_for_edits_id'] = message_id
            else:
                message_batches[user_id] = {
                    'messages': [user_text],
                    'first_update': update,
                    'waiting_for_edits_id': message_id,
                }
            task = asyncio.create_task(process_batched_messages(user_id, context))
            message_batches[user_id]['task'] = task
            return
        else:
            await update.message.reply_text("❌ Данные о генерации не найдены.")
            return

    # Determine if this is a reply to a generated post
    replied_message_id = None
    if update.message.reply_to_message:
        replied_message_id = update.message.reply_to_message.message_id
        # Only consider it a reply if generation data exists
        if not get_generation(context, replied_message_id):
            replied_message_id = None

    # Message batching logic - works for both regular messages and replies
    if user_id in message_batches:
        # Cancel existing timer and add message to batch
        batch = message_batches[user_id]
        batch['task'].cancel()
        batch['messages'].append(user_text)
        # Update replied_message_id if needed (should be same for all messages in batch)
        if replied_message_id:
            batch['replied_message_id'] = replied_message_id
        logger.info(f"Added message to batch for user {user_id}, total: {len(batch['messages'])}")
    else:
        # Create new batch
        batch_data = {
            'messages': [user_text],
            'first_update': update,
        }
        if replied_message_id:
            batch_data['replied_message_id'] = replied_message_id
        message_batches[user_id] = batch_data
        logger.info(f"Created new batch for user {user_id}" + (f" (reply to {replied_message_id})" if replied_message_id else ""))

    # Start new timer
    task = asyncio.create_task(process_batched_messages(user_id, context))
    message_batches[user_id]['task'] = task

async def start_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle /start command."""
    user = update.effective_user
    user_id = str(user.id)

    logger.info(f"Received /start command from {user.username} ({user_id})")

    welcome_text = """🤖 Привет! Я бот для создания вирусных постов из статей про стартапы.

<b>Как пользоваться:</b>
• Кинь мне ссылку на статью или текст
• Я сгенерирую пост в стиле Indie Hackers
• Ты сможешь переделать пост кнопками или ответом

<b>Как редактировать посты:</b>
• Нажми "✂️ Короче" для короткой версии
• Нажми "🔄 Переделать" для другого варианта
• Или просто ответь на пост с правками

<b>Батчинг сообщений:</b>
Если ты кидаешь длинный текст, который Telegram разбивает на части, я автоматически склею все части и обработаю как одно сообщение. Это работает и для ответов на посты!

Давай создадим что-то вирусное! 🚀"""

    await update.message.reply_text(welcome_text, parse_mode='HTML')

def format_story(story_data, index=None):
    """Format a single story for display."""
    prefix = f"{index}. " if index else "• "
    title = story_data.get('Title', 'No Title')
    author = story_data.get('Author', '')
    product = story_data.get('Product', '')
    mrr = story_data.get('MRR', '')
    url = story_data.get('URL', '')

    text = f"{prefix}<b>{title}</b>\n"
    if author:
        text += f"👤 {author}\n"
    if product:
        text += f"🚀 {product}\n"
    if mrr:
        text += f"💰 {mrr}\n"
    text += f"🔗 {url}\n"

    return text

def read_stories_from_excel(file_path):
    """Read stories from Excel file."""
    wb = load_workbook(file_path)
    ws = wb.active

    stories = []
    # Skip header row
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] or row[1]:  # At least ID or Title exists
            stories.append({
                'ID': row[0] or '',
                'Title': row[1] or '',
                'Author': row[2] or '',
                'Product': row[3] or '',
                'MRR': row[4] or '',
                'URL': row[5] or ''
            })

    return stories

def screenshot_mockup(input_path, output_path):
    """Transform screenshot into beautiful mockup with gradient background and macOS window."""
    from PIL import Image
    import io

    # Load image and convert to base64 PNG
    img = Image.open(input_path)

    # Convert to RGB if needed
    if img.mode in ('RGBA', 'LA'):
        background = Image.new('RGB', img.size, 'white')
        if img.mode == 'RGBA':
            background.paste(img, mask=img.split()[3])
        else:
            background.paste(img, mask=img.split()[1])
        img = background
    elif img.mode != 'RGB':
        img = img.convert('RGB')

    # Convert to base64
    buffer = io.BytesIO()
    img.save(buffer, format='PNG')
    img_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')

    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            * {{ margin: 0; padding: 0; box-sizing: border-box; }}
            body {{
                width: 100vw;
                height: 100vh;
                display: flex;
                align-items: center;
                justify-content: center;
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
            }}
            .window {{
                background: white;
                border-radius: 18px;
                box-shadow: 0 40px 120px rgba(0,0,0,0.35);
                overflow: hidden;
                padding-top: 32px;
                position: relative;
            }}
            .dots {{
                position: absolute;
                top: 10px;
                left: 14px;
                display: flex;
                gap: 8px;
            }}
            .dot {{
                width: 12px;
                height: 12px;
                border-radius: 50%;
            }}
            .dot.red {{ background: #ff5f57; }}
            .dot.yellow {{ background: #febc2e; }}
            .dot.green {{ background: #28c840; }}
            img {{
                display: block;
                max-width: 1100px;
                border-radius: 0 0 18px 18px;
            }}
        </style>
    </head>
    <body>
        <div class="window">
            <div class="dots">
                <span class="dot red"></span>
                <span class="dot yellow"></span>
                <span class="dot green"></span>
            </div>
            <img src="data:image/png;base64,{img_base64}" alt="Screenshot">
        </div>
    </body>
    </html>
    """

    # Get output directory and filename
    output_dir = Path(output_path).parent
    output_filename = Path(output_path).name

    # Create temp directory for html2image
    temp_dir = Path("photos/temp")
    temp_dir.mkdir(parents=True, exist_ok=True)

    # Create Html2Image with flags for running as root
    hti = Html2Image(
        output_path=str(output_dir),
        temp_path=str(temp_dir),  # Use photos/temp for temporary files
        size=(1600, 1000),
        browser='chromium',
        browser_executable='/usr/bin/chromium-browser',
        custom_flags=[
            '--no-sandbox',
            '--disable-dev-shm-usage',
            '--disable-gpu',
            '--disable-software-rasterizer'
        ]
    )
    hti.screenshot(html_str=html, save_as=output_filename)
    logger.info(f"Screenshot mockup created: {output_path}")

async def fetch_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    user_id = str(user.id)

    logger.info(f"Received /fetch command from {user.username} ({user_id})")

    # Filter by Admin ID if set
    if ADMIN_ID and user_id != str(ADMIN_ID):
        logger.warning(f"Unauthorized /fetch attempt by {user_id}")
        return

    status_message = await update.message.reply_text("⏳ Запускаю парсинг историй... (это может занять время)")

    try:
        # Run the fetch script using subprocess
        # Assumes the script is at scripts/fetch_stories.py relative to current dir
        script_path = Path("scripts/fetch_stories.py")
        result = subprocess.run(
            [sys.executable, str(script_path)],
            capture_output=True,
            text=True,
            check=True
        )

        logger.info(f"Fetch script output: {result.stdout}")

        # The script saves to output/stories.xlsx
        output_file = Path("output/stories.xlsx")

        if output_file.exists():
            await status_message.edit_text("✅ Парсинг завершен. Отправляю данные...")

            # Send the file
            with open(output_file, "rb") as f:
                await update.message.reply_document(
                    document=f,
                    filename="stories.xlsx",
                    caption="📊 Полный список историй"
                )

            # Read stories from the file
            stories = read_stories_from_excel(output_file)

            if stories:
                # Top 10 latest stories (first in the list)
                top_10 = stories[:10]
                top_10_text = "🔥 <b>Топ 10 последних историй:</b>\n\n"
                for i, story in enumerate(top_10, 1):
                    top_10_text += format_story(story, i) + "\n"

                # Send top 10 (may need to split if too long)
                try:
                    await update.message.reply_text(top_10_text, parse_mode='HTML', disable_web_page_preview=True)
                except Exception as e:
                    logger.warning(f"Failed to send top 10 as HTML: {e}")
                    # Try splitting or sending as plain text
                    await update.message.reply_text("⚠️ Список слишком длинный, смотри в файле")

                # 3 random stories
                if len(stories) > 10:
                    random_stories = random.sample(stories, min(3, len(stories)))
                    random_text = "🎲 <b>3 случайные истории:</b>\n\n"
                    for story in random_stories:
                        random_text += format_story(story) + "\n"

                    try:
                        await update.message.reply_text(random_text, parse_mode='HTML', disable_web_page_preview=True)
                    except Exception as e:
                        logger.warning(f"Failed to send random stories as HTML: {e}")
                        await update.message.reply_text("⚠️ Не удалось отправить случайные истории")

                await status_message.delete()
            else:
                await status_message.edit_text("⚠️ Файл пустой, истории не найдены")
        else:
            await status_message.edit_text("⚠️ Скрипт завершился, но файл не найден.")

    except subprocess.CalledProcessError as e:
        logger.error(f"Error running fetch script: {e.stderr}")
        await status_message.edit_text(f"❌ Ошибка при выполнении скрипта:\n{e.stderr[:200]}")
    except Exception as e:
        logger.error(f"Error in fetch_command: {e}", exc_info=True)
        await status_message.edit_text(f"❌ Неизвестная ошибка: {str(e)}")

async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle photo messages - download, apply mockup, and send back."""
    user = update.effective_user
    user_id = str(user.id)

    logger.info(f"Received photo from {user.username} ({user_id})")

    # Filter by Admin ID if set
    if ADMIN_ID and user_id != str(ADMIN_ID):
        logger.warning(f"Unauthorized photo from {user_id}")
        return

    status_message = await update.message.reply_text("🎨 Обрабатываю фото...")

    try:
        # Get the largest photo size
        photo = update.message.photo[-1]
        photo_file = await photo.get_file()

        # Create photos directory structure
        photos_base = Path("photos")
        photos_input = photos_base / "input"
        photos_output = photos_base / "output"

        # Create all directories
        for dir_path in [photos_input, photos_output]:
            dir_path.mkdir(parents=True, exist_ok=True)

        # Save in photos/input and photos/output
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        input_path = photos_input / f"{timestamp}.png"
        output_path = photos_output / f"{timestamp}.png"

        await photo_file.download_to_drive(input_path)
        logger.info(f"Downloaded photo to {input_path}")

        # Apply mockup in thread to not block
        await status_message.edit_text("✨ Применяю эффект...")
        await asyncio.to_thread(screenshot_mockup, str(input_path), str(output_path))

        # Send back the processed photo
        await status_message.edit_text("📤 Отправляю результат...")
        with open(output_path, "rb") as f:
            await update.message.reply_photo(
                photo=f,
                caption="✅ Готово! Красивый мокап с градиентом и macOS-окном 🎨"
            )

        # Delete status message
        await status_message.delete()

        logger.info(f"Photo processing completed: {output_path}")

    except Exception as e:
        logger.error(f"Error in handle_photo: {e}", exc_info=True)
        await status_message.edit_text(f"❌ Ошибка при обработке фото: {str(e)[:200]}")

async def random_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Fetch a random article from Indie Hackers or TechCrunch and generate a post."""
    user = update.effective_user
    user_id = str(user.id)

    logger.info(f"Received /random command from {user.username} ({user_id})")

    # Filter by Admin ID if set
    if ADMIN_ID and user_id != str(ADMIN_ID):
        logger.warning(f"Unauthorized /random attempt by {user_id}")
        return

    status_message = await update.message.reply_text("🎲 Ищу случайную статью...")

    try:
        # Randomly choose source: Indie Hackers or TechCrunch
        source = random.choice(['indiehackers', 'techcrunch'])

        article_url = None
        article_title = None
        article_info = None

        if source == 'techcrunch':
            # Fetch TechCrunch RSS
            logger.info("Fetching random article from TechCrunch")
            feed_url = 'https://techcrunch.com/feed/'
            feed = await asyncio.to_thread(feedparser.parse, feed_url)

            if not feed.entries:
                await status_message.edit_text("❌ Не удалось загрузить статьи TechCrunch")
                return

            # Get random article from first 20
            articles = feed.entries[:20]
            random_article = random.choice(articles)

            article_url = random_article.get('link', '')
            article_title = random_article.get('title', 'No Title')
            published = random_article.get('published', '')

            # Parse date
            date_str = ''
            if published:
                try:
                    from datetime import datetime
                    pub_date = datetime.strptime(published, '%a, %d %b %Y %H:%M:%S %z')
                    date_str = pub_date.strftime('%d.%m.%Y')
                except:
                    date_str = published.split(',')[0] if ',' in published else ''

            article_info = f"🔥 <b>TechCrunch</b>\n\n<b>{article_title}</b>\n"
            if date_str:
                article_info += f"📅 {date_str}\n"
            article_info += f"🔗 {article_url}\n\n⏳ Генерирую пост..."

        else:  # indiehackers
            # Read stories from Excel if exists
            logger.info("Fetching random article from Indie Hackers")
            output_file = Path("output/stories.xlsx")

            if not output_file.exists():
                # Fallback to TechCrunch if no stories file
                logger.warning("stories.xlsx not found, falling back to TechCrunch")
                await status_message.edit_text("⚠️ Файл с историями Indie Hackers не найден. Используйте /fetch сначала.\n\nПереключаюсь на TechCrunch...")
                await asyncio.sleep(2)

                # Fetch from TechCrunch instead
                feed_url = 'https://techcrunch.com/feed/'
                feed = await asyncio.to_thread(feedparser.parse, feed_url)

                if not feed.entries:
                    await status_message.edit_text("❌ Не удалось загрузить статьи")
                    return

                articles = feed.entries[:20]
                random_article = random.choice(articles)

                article_url = random_article.get('link', '')
                article_title = random_article.get('title', 'No Title')
                article_info = f"🔥 <b>TechCrunch</b>\n\n<b>{article_title}</b>\n🔗 {article_url}\n\n⏳ Генерирую пост..."
            else:
                stories = read_stories_from_excel(output_file)

                if not stories:
                    await status_message.edit_text("❌ Файл историй пуст")
                    return

                # Get random story from all stories (up to 100)
                stories_subset = stories[:min(100, len(stories))]
                random_story = random.choice(stories_subset)

                article_url = random_story.get('URL', '')
                article_title = random_story.get('Title', 'No Title')
                author = random_story.get('Author', '')
                product = random_story.get('Product', '')
                mrr = random_story.get('MRR', '')

                article_info = f"🚀 <b>Indie Hackers</b>\n\n<b>{article_title}</b>\n"
                if author:
                    article_info += f"👤 {author}\n"
                if product:
                    article_info += f"🏗️ {product}\n"
                if mrr:
                    article_info += f"💰 {mrr}\n"
                article_info += f"🔗 {article_url}\n\n⏳ Генерирую пост..."

        # Send article info
        await status_message.edit_text(article_info, parse_mode='HTML', disable_web_page_preview=True)

        # Now generate post from this URL
        if not article_url:
            await context.bot.send_message(chat_id=update.effective_chat.id, text="❌ Не удалось получить URL статьи")
            return

        logger.info(f"Generating post for random article: {article_url}")

        # Send typing action and new status message
        await context.bot.send_chat_action(chat_id=update.effective_chat.id, action=ChatAction.TYPING)
        generation_status = await context.bot.send_message(
            chat_id=update.effective_chat.id,
            text="🧠 Читаю статью...",
            reply_to_message_id=status_message.message_id
        )

        # Start status updater
        stop_event = asyncio.Event()
        status_task = asyncio.create_task(update_status_periodically(generation_status, stop_event))

        try:
            # Extract content
            content = extract_content(article_url)

            if content is None:
                stop_event.set()
                status_task.cancel()
                await generation_status.edit_text("❌ Не удалось извлечь контент из ссылки.")
                return

            system_prompt = get_prompt()

            # Save extracted content
            from datetime import datetime
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            debug_dir = Path("debug")
            debug_dir.mkdir(exist_ok=True)

            extracted_file = debug_dir / f"{user_id}_{timestamp}_random_extracted.md"
            with open(extracted_file, "w", encoding="utf-8") as f:
                f.write(content)
            logger.info(f"Saved extracted content to {extracted_file}")

            # Call LLM
            logger.info(f"Calling LLM ({LLM_MODEL})...")
            response = await asyncio.to_thread(
                completion,
                model=LLM_MODEL,
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": content}
                ],
                api_key=OPENROUTER_API_KEY
            )

            reply_text = response.choices[0].message.content
            logger.info(f"Successfully received response from LLM ({len(reply_text)} chars)")

            # Stop status updater
            stop_event.set()
            status_task.cancel()

            # Save LLM response
            response_file = debug_dir / f"{user_id}_{timestamp}_random_response.md"
            with open(response_file, "w", encoding="utf-8") as f:
                f.write(reply_text)
            logger.info(f"Saved LLM response to {response_file}")

            # Delete status message and send response
            await generation_status.delete()
            reply_text = strip_utm_params(reply_text)

            try:
                html_text = markdown_to_html(reply_text)
                sent_message = await context.bot.send_message(
                    chat_id=update.effective_chat.id,
                    text=html_text,
                    parse_mode='HTML',
                    reply_to_message_id=status_message.message_id,
                    reply_markup=create_post_keyboard(0)
                )
            except Exception as parse_error:
                logger.warning(f"HTML parse failed, sending as plain text: {parse_error}")
                sent_message = await context.bot.send_message(
                    chat_id=update.effective_chat.id,
                    text=reply_text,
                    reply_to_message_id=status_message.message_id,
                    reply_markup=create_post_keyboard(0)
                )

            # Update keyboard with correct message_id and save generation
            await sent_message.edit_reply_markup(reply_markup=create_post_keyboard(sent_message.message_id))
            save_generation(context, sent_message.message_id, content, reply_text)

        except Exception as e:
            stop_event.set()
            status_task.cancel()
            logger.error(f"Error generating post: {str(e)}", exc_info=True)
            await generation_status.edit_text(f"❌ Ошибка при генерации: {str(e)[:200]}")

    except Exception as e:
        logger.error(f"Error in random_command: {e}", exc_info=True)
        await status_message.edit_text(f"❌ Ошибка: {str(e)[:200]}")

async def techcrunch_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Fetch latest articles from TechCrunch RSS feed."""
    user = update.effective_user
    user_id = str(user.id)

    logger.info(f"Received /techcrunch command from {user.username} ({user_id})")

    # Filter by Admin ID if set
    if ADMIN_ID and user_id != str(ADMIN_ID):
        logger.warning(f"Unauthorized /techcrunch attempt by {user_id}")
        return

    # Parse arguments: /techcrunch [count] [category]
    args = context.args
    count = 10  # default
    category = None

    # Available categories
    categories = {
        'startups': 'https://techcrunch.com/category/startups/feed/',
        'ai': 'https://techcrunch.com/category/artificial-intelligence/feed/',
        'apps': 'https://techcrunch.com/category/apps/feed/',
        'crypto': 'https://techcrunch.com/category/cryptocurrency/feed/',
        'venture': 'https://techcrunch.com/category/venture/feed/',
        'security': 'https://techcrunch.com/category/security/feed/',
    }

    if args:
        # Check if first arg is a number
        if args[0].isdigit():
            count = min(int(args[0]), 50)  # max 50
            if len(args) > 1:
                category = args[1].lower()
        else:
            category = args[0].lower()

    # Determine feed URL
    if category and category in categories:
        feed_url = categories[category]
        category_name = category.capitalize()
    else:
        feed_url = 'https://techcrunch.com/feed/'
        category_name = 'All'

    status_message = await update.message.reply_text(f"⏳ Загружаю статьи TechCrunch ({category_name})...")

    try:
        # Fetch RSS feed
        logger.info(f"Fetching TechCrunch RSS: {feed_url}")
        feed = await asyncio.to_thread(feedparser.parse, feed_url)

        if not feed.entries:
            await status_message.edit_text("❌ Не удалось загрузить статьи из RSS-ленты")
            return

        # Get requested number of entries
        entries = feed.entries[:count]

        # Format articles
        response_text = f"🔥 <b>TechCrunch - {category_name}</b>\n"
        response_text += f"Последние {len(entries)} статей:\n\n"

        for i, entry in enumerate(entries, 1):
            title = entry.get('title', 'No Title')
            link = entry.get('link', '')
            published = entry.get('published', '')

            # Parse date if available
            date_str = ''
            if published:
                try:
                    from datetime import datetime
                    pub_date = datetime.strptime(published, '%a, %d %b %Y %H:%M:%S %z')
                    date_str = pub_date.strftime('%d.%m.%Y')
                except:
                    date_str = published.split(',')[0] if ',' in published else ''

            response_text += f"{i}. <b>{title}</b>\n"
            if date_str:
                response_text += f"📅 {date_str}\n"
            response_text += f"🔗 {link}\n\n"

        # Add category help
        if not category:
            response_text += "\n💡 <b>Доступные категории:</b>\n"
            response_text += "• /techcrunch 10 startups\n"
            response_text += "• /techcrunch ai\n"
            response_text += "• /techcrunch crypto\n"
            response_text += "• /techcrunch venture\n"
            response_text += "• /techcrunch security\n"

        response_text += "\n📝 Кинь любую ссылку в чат для генерации поста!"

        await status_message.delete()

        # Send in chunks if too long (Telegram limit is 4096)
        if len(response_text) > 4000:
            # Split by articles
            chunks = []
            current_chunk = f"🔥 <b>TechCrunch - {category_name}</b>\n"
            current_chunk += f"Последние {len(entries)} статей:\n\n"

            for i, entry in enumerate(entries, 1):
                title = entry.get('title', 'No Title')
                link = entry.get('link', '')
                published = entry.get('published', '')

                date_str = ''
                if published:
                    try:
                        from datetime import datetime
                        pub_date = datetime.strptime(published, '%a, %d %b %Y %H:%M:%S %z')
                        date_str = pub_date.strftime('%d.%m.%Y')
                    except:
                        date_str = published.split(',')[0] if ',' in published else ''

                article_text = f"{i}. <b>{title}</b>\n"
                if date_str:
                    article_text += f"📅 {date_str}\n"
                article_text += f"🔗 {link}\n\n"

                if len(current_chunk) + len(article_text) > 3900:
                    chunks.append(current_chunk)
                    current_chunk = ""

                current_chunk += article_text

            if current_chunk:
                chunks.append(current_chunk)

            # Send chunks
            for chunk in chunks:
                await update.message.reply_text(chunk, parse_mode='HTML', disable_web_page_preview=True)
        else:
            await update.message.reply_text(response_text, parse_mode='HTML', disable_web_page_preview=True)

        logger.info(f"Successfully sent {len(entries)} TechCrunch articles")

    except Exception as e:
        logger.error(f"Error in techcrunch_command: {e}", exc_info=True)
        await status_message.edit_text(f"❌ Ошибка при загрузке TechCrunch: {str(e)[:200]}")

async def scheduled_random_post(app):
    """Send a random post at scheduled time (8 AM MSK)."""
    logger.info("Running scheduled random post task")

    if not ADMIN_ID:
        logger.warning("ADMIN_ID not set, skipping scheduled post")
        return

    try:
        chat_id = int(ADMIN_ID)

        # Send initial message
        status_message = await app.bot.send_message(
            chat_id=chat_id,
            text="🌅 Доброе утро! 🎲 Ищу случайную статью для тебя..."
        )

        # Choose random source
        source = random.choice(['indiehackers', 'techcrunch'])

        article_url = None
        article_title = None
        article_info = None

        if source == 'techcrunch':
            logger.info("Scheduled task: Fetching from TechCrunch")
            feed_url = 'https://techcrunch.com/feed/'
            feed = await asyncio.to_thread(feedparser.parse, feed_url)

            if not feed.entries:
                await status_message.edit_text("❌ Не удалось загрузить статьи TechCrunch")
                return

            articles = feed.entries[:20]
            random_article = random.choice(articles)

            article_url = random_article.get('link', '')
            article_title = random_article.get('title', 'No Title')
            published = random_article.get('published', '')

            date_str = ''
            if published:
                try:
                    from datetime import datetime
                    pub_date = datetime.strptime(published, '%a, %d %b %Y %H:%M:%S %z')
                    date_str = pub_date.strftime('%d.%m.%Y')
                except:
                    date_str = published.split(',')[0] if ',' in published else ''

            article_info = f"🔥 <b>TechCrunch</b>\n\n<b>{article_title}</b>\n"
            if date_str:
                article_info += f"📅 {date_str}\n"
            article_info += f"🔗 {article_url}\n\n⏳ Генерирую пост..."

        else:  # indiehackers
            logger.info("Scheduled task: Fetching from Indie Hackers")
            output_file = Path("output/stories.xlsx")

            if not output_file.exists():
                logger.warning("Scheduled task: stories.xlsx not found, using TechCrunch")
                await status_message.edit_text("⚠️ Файл с историями не найден, переключаюсь на TechCrunch...")
                await asyncio.sleep(2)

                feed_url = 'https://techcrunch.com/feed/'
                feed = await asyncio.to_thread(feedparser.parse, feed_url)

                if not feed.entries:
                    await status_message.edit_text("❌ Не удалось загрузить статьи")
                    return

                articles = feed.entries[:20]
                random_article = random.choice(articles)

                article_url = random_article.get('link', '')
                article_title = random_article.get('title', 'No Title')
                article_info = f"🔥 <b>TechCrunch</b>\n\n<b>{article_title}</b>\n🔗 {article_url}\n\n⏳ Генерирую пост..."
            else:
                stories = read_stories_from_excel(output_file)

                if not stories:
                    await status_message.edit_text("❌ Файл историй пуст")
                    return

                stories_subset = stories[:min(100, len(stories))]
                random_story = random.choice(stories_subset)

                article_url = random_story.get('URL', '')
                article_title = random_story.get('Title', 'No Title')
                author = random_story.get('Author', '')
                product = random_story.get('Product', '')
                mrr = random_story.get('MRR', '')

                article_info = f"🚀 <b>Indie Hackers</b>\n\n<b>{article_title}</b>\n"
                if author:
                    article_info += f"👤 {author}\n"
                if product:
                    article_info += f"🏗️ {product}\n"
                if mrr:
                    article_info += f"💰 {mrr}\n"
                article_info += f"🔗 {article_url}\n\n⏳ Генерирую пост..."

        # Send article info
        await status_message.edit_text(article_info, parse_mode='HTML', disable_web_page_preview=True)

        if not article_url:
            await app.bot.send_message(chat_id=chat_id, text="❌ Не удалось получить URL статьи")
            return

        logger.info(f"Scheduled task: Generating post for {article_url}")

        # Send typing and status
        await app.bot.send_chat_action(chat_id=chat_id, action=ChatAction.TYPING)
        generation_status = await app.bot.send_message(
            chat_id=chat_id,
            text="🧠 Читаю статью...",
            reply_to_message_id=status_message.message_id
        )

        stop_event = asyncio.Event()
        status_task = asyncio.create_task(update_status_periodically(generation_status, stop_event))

        try:
            content = extract_content(article_url)

            if content is None:
                stop_event.set()
                status_task.cancel()
                await generation_status.edit_text("❌ Не удалось извлечь контент")
                return

            system_prompt = get_prompt()

            # Call LLM
            logger.info(f"Scheduled task: Calling LLM...")
            response = await asyncio.to_thread(
                completion,
                model=LLM_MODEL,
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": content}
                ],
                api_key=OPENROUTER_API_KEY
            )

            reply_text = response.choices[0].message.content
            logger.info(f"Scheduled task: Received LLM response ({len(reply_text)} chars)")

            stop_event.set()
            status_task.cancel()

            await generation_status.delete()
            reply_text = strip_utm_params(reply_text)

            try:
                html_text = markdown_to_html(reply_text)
                sent_message = await app.bot.send_message(
                    chat_id=chat_id,
                    text=html_text,
                    parse_mode='HTML',
                    reply_to_message_id=status_message.message_id,
                    reply_markup=create_post_keyboard(0)
                )
            except Exception as parse_error:
                logger.warning(f"HTML parse failed: {parse_error}")
                sent_message = await app.bot.send_message(
                    chat_id=chat_id,
                    text=reply_text,
                    reply_to_message_id=status_message.message_id,
                    reply_markup=create_post_keyboard(0)
                )

            # Update keyboard
            await sent_message.edit_reply_markup(reply_markup=create_post_keyboard(sent_message.message_id))

            logger.info("Scheduled task: Successfully sent random post")

        except Exception as e:
            stop_event.set()
            status_task.cancel()
            logger.error(f"Error in scheduled post generation: {str(e)}", exc_info=True)
            await generation_status.edit_text(f"❌ Ошибка: {str(e)[:200]}")

    except Exception as e:
        logger.error(f"Error in scheduled_random_post: {e}", exc_info=True)

async def post_init(app):
    """Initialize scheduler after bot starts."""
    scheduler = AsyncIOScheduler(timezone=pytz.timezone('Europe/Moscow'))

    # Schedule daily post at 8:00 AM MSK
    scheduler.add_job(
        scheduled_random_post,
        trigger=CronTrigger(hour=8, minute=0, timezone=pytz.timezone('Europe/Moscow')),
        args=[app],
        id='daily_morning_post',
        name='Daily morning random post at 8 AM MSK',
        replace_existing=True
    )

    scheduler.start()
    logger.info("Scheduler started - Daily post will be sent at 8:00 AM MSK")

if __name__ == '__main__':
    if not TELEGRAM_BOT_TOKEN:
        logger.error("Error: TELEGRAM_BOT_TOKEN not found in .env")
        exit(1)

    app = ApplicationBuilder().token(TELEGRAM_BOT_TOKEN).post_init(post_init).build()

    app.add_handler(CommandHandler("start", start_command))
    app.add_handler(CommandHandler("fetch", fetch_command))
    app.add_handler(CommandHandler("techcrunch", techcrunch_command))
    app.add_handler(CommandHandler("random", random_command))
    app.add_handler(CallbackQueryHandler(handle_callback))
    app.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    app.add_handler(MessageHandler(filters.TEXT & (~filters.COMMAND), handle_message))

    logger.info(f"Bot is starting... (Admin: {ADMIN_ID if ADMIN_ID else 'None'})")
    app.run_polling()
